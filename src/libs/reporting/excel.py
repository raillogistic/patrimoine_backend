import platform
import tempfile

from django.http import HttpResponse
from openpyxl import Workbook, utils
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side

from libs.utils.tempfile import write_doc

can_delete = False if platform.system() == "Windows" else True


def set_auto_filter(worksheet, column_count):
    """Set auto filter for the header row."""
    worksheet.auto_filter.ref = (
        worksheet.dimensions
    )  # Apply filter to the full range of data


def set_auto_sum(
    self,
    worksheet,
):
    """Set sum of each column."""
    max_row = worksheet.max_row
    max_col = worksheet.max_column

    total_row = max_row + 1  # The row for the totals, right after the last data row
    for col in range(1, max_col + 1):
        column_letter = utils.get_column_letter(col)
        cell = worksheet[f"{column_letter}{total_row}"]

        # Check if the column has numeric data, and if so, add a sum formula
        if all(
            isinstance(worksheet[f"{column_letter}{row}"].value, (int, float))
            for row in range(2, max_row + 1)
        ):
            print(
                "xxxxxxxx",
            )
            cell.value = f"=SUM({column_letter}2:{column_letter}{max_row})"
            cell.number_format = (
                "#,##0.00"  # Apply number formatting for the total cell
            )
            cell.font = Font(bold=True)  # Make the total cell bold


def apply_borders(worksheet):
    """Apply borders to all cells in the worksheet."""
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = thin_border


def apply_background(worksheet):
    """Apply borders to all cells in the worksheet."""

    for row in worksheet.iter_rows():
        for cell in row:
            if cell.row % 2 == 0:
                cell.fill = PatternFill(
                    start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
                )  # Blue background
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"


# cell.column % 2 == 0
def adjust_column_widths(worksheet):
    """Adjust the width of the columns to fit the content."""
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter  # Get the column letter

        for cell in column:
            try:
                if len(str(cell.value.upper())) > max_length:
                    max_length = len(str(cell.value.upper()))
            except:
                pass

        # Set the column width with some extra space
        adjusted_width = max_length + 10
        worksheet.column_dimensions[column_letter].width = adjusted_width


def style_header_row(worksheet):
    """Style the header row of the worksheet."""
    header_row = worksheet[1]  # The first row for headers

    for cell in header_row:
        cell.value = cell.value.upper()
        cell.font = Font(bold=True, color="ffffff")  # Bold, white font
        cell.fill = PatternFill(
            start_color="0070C0", end_color="0070C0", fill_type="solid"
        )  # Blue background
        cell.alignment = Alignment(horizontal="left", vertical="center")
        border_color = "000000"
        cell.border = Border(
            left=Side(style="thin", color=border_color),
            right=Side(style="thin", color=border_color),
            top=Side(style="thin", color=border_color),
            bottom=Side(style="thin", color=border_color),
        )
    worksheet.row_dimensions[1].height = 20


def fix_currency(worksheet):
    currency_format = NamedStyle(name="currency")
    currency_format.number_format = "#,##0.00"

    # Apply background color to even columns and number format for currency
    print(worksheet)
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            # Apply currency format to 'Salary' column (assuming it's column 4)
            if isinstance(cell.value, (int, float)):
                cell.style = currency_format


def apply_styling(
    worksheet,
):
    fix_currency(worksheet)
    adjust_column_widths(worksheet)
    # Apply borders to all cell
    apply_borders(worksheet)
    style_header_row(worksheet)
    apply_background(worksheet)
    # Set filters for headers


def to_excel_file(filename, headers=[], data=[[]], title=""):
    """
    Django view to create and download an Excel file using a temporary file.
    """
    # Data to write into the Excel file

    # Create a workbook and a worksheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    # Write data to the worksheet
    for row in data:
        sheet.append(row)

    # apply_styling(workbook)

    # Use a temporary file to save the Excel workbook
    with tempfile.NamedTemporaryFile(delete=can_delete, suffix=".xlsx") as tmp_file:
        workbook.save(tmp_file.name)  # Save the workbook to the temporary file
        tmp_file.seek(0)  # Move to the start of the file

        # Prepare the HTTP response with the Excel file
        response = HttpResponse(
            tmp_file.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'

    # Temporary file is automatically deleted when closed
    return response
